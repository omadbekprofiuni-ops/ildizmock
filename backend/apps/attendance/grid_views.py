"""ETAP 20 — Attendance Redesign 3-tab API.

Endpoints (mounted at `/api/v1/center/<slug>/attendance/`):
- GET  v2/today-session/             — Tab 1: bugungi/tanlangan sessiya + records
- POST v2/sessions/<id>/bulk-mark/   — Tab 1: barcha talabalarni saqlash
- GET  v2/monthly-grid/              — Tab 2: oylik jadval
- PATCH v2/records/<id>/             — Tab 2: bitta hujayrani belgilash
- GET  v2/analytics/                 — Tab 3: trend + at_risk
- GET  v2/students/<id>/history/     — Talaba detail sahifa
- GET  v2/export.xlsx                — Tab 2: Excel eksport
"""
from __future__ import annotations

from datetime import date as date_cls
from datetime import datetime

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import permissions, status as drf_status
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.organizations.models import (
    Organization,
    OrganizationMembership,
    StudentGroup,
)

from .models import AttendanceRecord, AttendanceSession
from .services import grid as grid_svc

User = get_user_model()


# ===== Permissions / helpers =====


def _get_org_or_403(user, slug: str) -> Organization:
    org = get_object_or_404(Organization, slug=slug)
    if org.status != 'active':
        raise PermissionDenied('Center is not active.')
    if user.role == 'superadmin':
        return org
    is_admin = OrganizationMembership.objects.filter(
        user=user, organization=org, role__in=['admin', 'owner'],
    ).exists()
    if is_admin:
        return org
    if user.role == 'teacher' and user.organization_id == org.id:
        return org
    raise PermissionDenied("Siz bu markaz a'zosi emassiz.")


def _is_admin(user, org: Organization) -> bool:
    if user.role == 'superadmin':
        return True
    return OrganizationMembership.objects.filter(
        user=user, organization=org, role__in=['admin', 'owner'],
    ).exists()


def _accessible_groups(user, org: Organization):
    """Admin — barcha guruhlar, teacher — faqat o'zining."""
    qs = StudentGroup.objects.filter(organization=org, is_active=True)
    if not _is_admin(user, org):
        qs = qs.filter(teacher=user)
    return qs


def _group_or_403(user, org: Organization, group_id) -> StudentGroup:
    try:
        return _accessible_groups(user, org).get(pk=group_id)
    except StudentGroup.DoesNotExist:
        raise PermissionDenied('Group not accessible.')


def _parse_date(value: str | None) -> date_cls | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def _student_payload(user_obj) -> dict:
    return {
        'id': user_obj.id,
        'name': f'{user_obj.first_name} {user_obj.last_name}'.strip() or user_obj.username,
        'username': user_obj.username,
        'photo_url': None,  # avatar field hozircha yo'q
    }


# ===== Tab 1 — Today =====


class TodaySessionView(APIView):
    """`GET /attendance/v2/today-session/?group=<id>&date=YYYY-MM-DD`."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, org_slug: str):
        org = _get_org_or_403(request.user, org_slug)
        group_id = request.query_params.get('group')
        if not group_id:
            return Response({'error': 'group is required'}, status=400)
        group = _group_or_403(request.user, org, group_id)

        target_date = _parse_date(request.query_params.get('date')) or date_cls.today()
        session = grid_svc.get_today_session(group, requested_date=target_date)

        if not session:
            # Sessiya yo'q bo'lsa default talabalar ro'yxati qaytariladi
            students = list(group.members.filter(role='student', is_active=True).order_by(
                'first_name', 'last_name',
            ))
            return Response({
                'session': None,
                'date': target_date.isoformat(),
                'group': {'id': group.id, 'name': group.name},
                'students': [_student_payload(s) for s in students],
                'records': [],
            })

        records = list(
            session.records.select_related('student').order_by(
                'student__first_name', 'student__last_name',
            )
        )
        return Response({
            'date': target_date.isoformat(),
            'group': {'id': group.id, 'name': group.name},
            'session': {
                'id': session.id,
                'date': session.date.isoformat(),
                'start_time': session.start_time.strftime('%H:%M') if session.start_time else None,
                'end_time': session.end_time.strftime('%H:%M') if session.end_time else None,
                'is_finalized': session.is_finalized,
                'is_locked': session.is_locked(),
                'notes': session.notes,
            },
            'records': [
                {
                    'id': r.id,
                    'student_id': r.student_id,
                    'student_name': f'{r.student.first_name} {r.student.last_name}'.strip()
                    or r.student.username,
                    'username': r.student.username,
                    'photo_url': None,
                    'status': r.status,
                    'note': r.notes or '',
                }
                for r in records
            ],
        })


class TodaySessionBulkMarkView(APIView):
    """`POST /attendance/v2/sessions/<id>/bulk-mark/`.

    Body: `{records: [{record_id?, student_id?, status, note}, ...]}`
    """

    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, org_slug: str, session_id: int):
        org = _get_org_or_403(request.user, org_slug)
        session = get_object_or_404(AttendanceSession, pk=session_id)
        _group_or_403(request.user, org, session.group_id)

        if session.is_locked():
            return Response({'error': 'session_locked'}, status=423)

        items = request.data.get('records') or []
        if not isinstance(items, list):
            return Response({'error': "records must be a list"}, status=400)

        try:
            saved = grid_svc.bulk_mark(session, items, marked_by=request.user)
        except ValidationError as e:
            return Response({'error': e.message_dict if hasattr(e, 'message_dict') else str(e)}, status=400)

        return Response({
            'saved': saved,
            'attendance_rate': session.get_attendance_rate(),
        })


# ===== Tab 2 — Monthly grid =====


class MonthlyGridView(APIView):
    """`GET /attendance/v2/monthly-grid/?group=<id>&year=YYYY&month=MM`."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, org_slug: str):
        org = _get_org_or_403(request.user, org_slug)
        group_id = request.query_params.get('group')
        if not group_id:
            return Response({'error': 'group is required'}, status=400)
        group = _group_or_403(request.user, org, group_id)

        today = date_cls.today()
        try:
            year = int(request.query_params.get('year', today.year))
            month = int(request.query_params.get('month', today.month))
        except (TypeError, ValueError):
            return Response({'error': 'invalid year/month'}, status=400)
        if not (1 <= month <= 12):
            return Response({'error': 'invalid month'}, status=400)

        data = grid_svc.get_monthly_grid(group, year, month)

        return Response({
            'group': {'id': group.id, 'name': group.name},
            'year': year,
            'month': month,
            'students': [_student_payload(s) for s in data['students']],
            'sessions': [
                {
                    'id': s.id,
                    'date': s.date.isoformat(),
                    'day': s.date.day,
                    'weekday': s.date.strftime('%a'),
                    'is_locked': s.is_locked(),
                }
                for s in data['sessions']
            ],
            'cells': [
                {'student_id': sid, 'session_id': ssid, **payload}
                for (sid, ssid), payload in data['cells'].items()
            ],
            'summary': data['summary'],
            'group_stats': data['group_stats'],
        })


class CellUpdateView(APIView):
    """`PATCH /attendance/v2/records/<id>/`.

    Body: `{status, note}`
    """

    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, org_slug: str, record_id: int):
        org = _get_org_or_403(request.user, org_slug)
        record = get_object_or_404(
            AttendanceRecord.objects.select_related('session', 'session__group'),
            pk=record_id,
        )
        _group_or_403(request.user, org, record.session.group_id)
        if record.session.is_locked():
            return Response({'error': 'session_locked'}, status=423)

        status_val = request.data.get('status')
        if not status_val:
            return Response({'error': 'status is required'}, status=400)
        if status_val not in dict(AttendanceRecord.STATUS_CHOICES):
            return Response({'error': 'invalid status'}, status=400)

        try:
            grid_svc.mark_single(
                record,
                status=status_val,
                note=request.data.get('note', ''),
                marked_by=request.user,
            )
        except ValidationError as e:
            return Response(
                {'error': e.message_dict if hasattr(e, 'message_dict') else str(e)},
                status=400,
            )

        return Response({
            'id': record.id,
            'status': record.status,
            'note': record.notes,
        })


# ===== Tab 3 — Analytics =====


class GridAnalyticsView(APIView):
    """`GET /attendance/v2/analytics/?group=<id>` (group ixtiyoriy).

    Group berilsa: shu guruh trendi + guruh ichidagi at_risk.
    Group berilmasa: markazning umumiy at_risk.
    """

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, org_slug: str):
        org = _get_org_or_403(request.user, org_slug)
        group_id = request.query_params.get('group')

        response: dict = {}

        if group_id:
            group = _group_or_403(request.user, org, group_id)
            response['group'] = {'id': group.id, 'name': group.name}
            response['trend'] = grid_svc.get_attendance_trend(group, months_back=6)
            org_risk = grid_svc.get_at_risk_students(org, threshold=70, months_back=3)
            # Faqat shu guruhdagi at_risk talabalarni filterlash.
            group_student_ids = set(
                group.members.filter(role='student').values_list('id', flat=True)
            )
            response['at_risk'] = [r for r in org_risk if r['student_id'] in group_student_ids]
        else:
            response['at_risk'] = grid_svc.get_at_risk_students(org, threshold=70, months_back=3)[:30]

        return Response(response)


# ===== Student detail =====


class StudentHistoryView(APIView):
    """`GET /attendance/v2/students/<student_id>/history/`."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, org_slug: str, student_id: int):
        org = _get_org_or_403(request.user, org_slug)
        student = get_object_or_404(
            User, pk=student_id, organization=org, role='student',
        )
        # Teacher faqat o'z guruhi talabasini ko'radi
        if not _is_admin(request.user, org):
            if student.group is None or student.group.teacher_id != request.user.id:
                raise PermissionDenied('Only students in your own group.')

        data = grid_svc.get_student_history(student, limit=50)

        return Response({
            'student': {
                'id': student.id,
                'name': f'{student.first_name} {student.last_name}'.strip() or student.username,
                'username': student.username,
                'photo_url': None,
                'group_name': student.group.name if student.group_id else None,
            },
            'overall_percent': data['overall_percent'],
            'total_marked': data['total_marked'],
            'by_status': data['by_status'],
            'records': [
                {
                    'id': r.id,
                    'session_id': r.session_id,
                    'session_date': r.session.date.isoformat(),
                    'group_name': r.session.group.name,
                    'status': r.status,
                    'note': r.notes or '',
                }
                for r in data['records']
            ],
        })


# ===== Excel export =====


class GridExcelExportView(APIView):
    """`GET /attendance/v2/export.xlsx?group=<id>&year=YYYY&month=MM`."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, org_slug: str):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        org = _get_org_or_403(request.user, org_slug)
        group_id = request.query_params.get('group')
        if not group_id:
            return Response({'error': 'group is required'}, status=400)
        group = _group_or_403(request.user, org, group_id)

        today = date_cls.today()
        try:
            year = int(request.query_params.get('year', today.year))
            month = int(request.query_params.get('month', today.month))
        except (TypeError, ValueError):
            return Response({'error': 'invalid year/month'}, status=400)

        data = grid_svc.get_monthly_grid(group, year, month)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'{year}-{month:02d}'[:31]

        ws.cell(1, 1, f'Davomat — {group.name}').font = Font(bold=True, size=14)
        ws.cell(2, 1, f'Oy: {year}-{month:02d}')

        row = 4
        ws.cell(row, 1, '№')
        ws.cell(row, 2, 'F.I.O')
        for col, sess in enumerate(data['sessions'], start=3):
            ws.cell(row, col, sess.date.day)
        last_col = 3 + len(data['sessions'])
        ws.cell(row, last_col, 'Davomat %')
        ws.cell(row, last_col + 1, 'Keldi')
        ws.cell(row, last_col + 2, 'Kech')
        ws.cell(row, last_col + 3, 'Kelmadi')
        ws.cell(row, last_col + 4, 'Sababli')
        ws.cell(row, last_col + 5, 'Jami')

        for c in range(1, last_col + 6):
            ws.cell(row, c).font = Font(bold=True)
            ws.cell(row, c).alignment = Alignment(horizontal='center')

        green = PatternFill('solid', fgColor='C6EFCE')
        yellow = PatternFill('solid', fgColor='FFEB9C')
        red = PatternFill('solid', fgColor='FFC7CE')
        gray = PatternFill('solid', fgColor='D9D9D9')

        row += 1
        for idx, s in enumerate(data['students'], start=1):
            ws.cell(row, 1, idx).alignment = Alignment(horizontal='center')
            ws.cell(row, 2, f'{s.first_name} {s.last_name}'.strip() or s.username)
            for col, sess in enumerate(data['sessions'], start=3):
                cell_data = data['cells'].get((s.id, sess.id), {})
                status_ = cell_data.get('status')
                cell = ws.cell(row, col, '')
                if status_ == 'present':
                    cell.value = '✓'
                    cell.fill = green
                elif status_ == 'late':
                    cell.value = '⏱'
                    cell.fill = yellow
                elif status_ == 'absent':
                    cell.value = '✗'
                    cell.fill = red
                elif status_ in ('excused', 'sick'):
                    cell.value = 'E'
                    cell.fill = gray
                cell.alignment = Alignment(horizontal='center')

            summary = data['summary'][s.id]
            ws.cell(row, last_col, summary['percent'] if summary['percent'] is not None else '—')
            ws.cell(row, last_col + 1, summary['present'])
            ws.cell(row, last_col + 2, summary['late'])
            ws.cell(row, last_col + 3, summary['absent'])
            ws.cell(row, last_col + 4, summary['excused'] + summary['sick'])
            ws.cell(row, last_col + 5, summary['total_sessions'])
            row += 1

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        for col in range(3, last_col):
            ws.column_dimensions[get_column_letter(col)].width = 5
        for col in range(last_col, last_col + 6):
            ws.column_dimensions[get_column_letter(col)].width = 10

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        safe_name = group.name.replace(' ', '_').replace('/', '_')
        response['Content-Disposition'] = (
            f'attachment; filename="davomat_{safe_name}_{year}_{month:02d}.xlsx"'
        )
        wb.save(response)
        return response
