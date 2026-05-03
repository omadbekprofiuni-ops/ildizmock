"""Bulk-import a Reading test from an Excel/.xlsx file.

The same file shape is used by the template-download endpoint and the
import endpoint. One row per question; passages are derived by grouping
on the "Passage Title" column.
"""
from __future__ import annotations

import io
import logging

import openpyxl
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.decorators import (
    api_view,
    authentication_classes,
    permission_classes,
)
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.accounts.authentication import CookieJWTAuthentication
from apps.organizations.models import Organization, OrganizationMembership
from apps.organizations.permissions import IsCenterAdmin
from apps.tests.models import Passage, Question, Test

log = logging.getLogger(__name__)

REQUIRED_COLUMNS = [
    'Passage Title',
    'Passage Content',
    'Question Number',
    'Question Text',
    'Question Type',
    'Correct Answer',
]
OPTION_COLUMNS = ['Option A', 'Option B', 'Option C', 'Option D']
VALID_TYPES = {'mcq', 'tfng', 'fill', 'short_answer'}

SAMPLE_ROWS = [
    {
        'Passage Title': 'The Origins of Coffee',
        'Passage Content': (
            "Coffee is one of the world's most popular beverages, "
            "with an estimated 2.25 billion cups consumed daily. "
            "Its origins trace back to the Ethiopian plateau, where "
            "legend tells of a goat herder named Kaldi who first noticed "
            'the energizing effects of the coffee plant. From Ethiopia, '
            "coffee spread to the Arabian Peninsula by the 15th century, "
            "where it was first cultivated and traded. By the 17th century, "
            "coffee had reached Europe and quickly became popular."
        ),
        'Question Number': 1,
        'Question Text': 'Coffee was first discovered in Ethiopia.',
        'Question Type': 'tfng',
        'Correct Answer': 'TRUE',
        'Option A': '',
        'Option B': '',
        'Option C': '',
        'Option D': '',
    },
    {
        'Passage Title': 'The Origins of Coffee',
        'Passage Content': '(same as above — leave blank or repeat)',
        'Question Number': 2,
        'Question Text': 'According to the legend, who first noticed coffee\'s effects?',
        'Question Type': 'mcq',
        'Correct Answer': 'A goat herder named Kaldi',
        'Option A': 'A goat herder named Kaldi',
        'Option B': 'An Arabian trader',
        'Option C': 'A European explorer',
        'Option D': 'An Ethiopian king',
    },
    {
        'Passage Title': 'The Origins of Coffee',
        'Passage Content': '(same as above)',
        'Question Number': 3,
        'Question Text': 'Coffee reached Europe in the ___ century.',
        'Question Type': 'fill',
        'Correct Answer': '17th',
        'Option A': '',
        'Option B': '',
        'Option C': '',
        'Option D': '',
    },
]


def _check_center_admin(user, slug: str) -> Organization:
    org = get_object_or_404(Organization, slug=slug)
    if user.role == 'superadmin':
        return org
    if not OrganizationMembership.objects.filter(
        user=user, organization=org, role__in=['admin', 'owner'],
    ).exists():
        from rest_framework.exceptions import PermissionDenied
        raise PermissionDenied('You are not an admin of this center.')
    return org


@api_view(['GET'])
@authentication_classes([CookieJWTAuthentication])
@permission_classes([IsAuthenticated, IsCenterAdmin])
def excel_template(request, org_slug=None):
    """Stream an .xlsx template with sample rows so the user knows the format."""
    _check_center_admin(request.user, org_slug)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Test'

    headers = REQUIRED_COLUMNS + OPTION_COLUMNS
    ws.append(headers)
    # Bold header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for row in SAMPLE_ROWS:
        ws.append([row.get(h, '') for h in headers])

    # Reasonable column widths
    widths = {'A': 28, 'B': 60, 'C': 10, 'D': 50, 'E': 14, 'F': 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for col in ('G', 'H', 'I', 'J'):
        ws.column_dimensions[col].width = 22

    # Add an instructions sheet
    ws2 = wb.create_sheet('Instructions')
    ws2.append(['Column', 'Required?', 'Notes'])
    for cell in ws2[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    rows = [
        ('Passage Title', 'Yes', 'Group rows by passage. Repeat the title for each row in the same passage.'),
        ('Passage Content', 'Yes (first row of each passage)', 'Full passage text. Subsequent rows can leave this blank.'),
        ('Question Number', 'Yes', 'Integer ordering within the test (1, 2, 3, …).'),
        ('Question Text', 'Yes', 'The question prompt shown to the student.'),
        ('Question Type', 'Yes', 'One of: mcq, tfng, fill, short_answer'),
        ('Correct Answer', 'Yes', 'For tfng: TRUE / FALSE / NOT GIVEN. For mcq: the exact text of the correct option.'),
        ('Option A–D', 'Only for mcq', 'Leave blank for non-mcq questions.'),
    ]
    for r in rows:
        ws2.append(r)
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="ielts_test_template.xlsx"'
    return resp


def _read_rows(workbook):
    """Yield {column: value} dicts from the first sheet, skipping the header row."""
    ws = workbook.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return [], []
    headers = [str(h).strip() if h is not None else '' for h in headers]

    out = []
    for row in rows_iter:
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        out.append({headers[i]: row[i] for i in range(len(headers))})
    return headers, out


@api_view(['POST'])
@authentication_classes([CookieJWTAuthentication])
@permission_classes([IsAuthenticated, IsCenterAdmin])
def excel_import(request, org_slug=None):
    """POST <slug>/tests/excel-import/

    multipart/form-data:
      file: .xlsx file (required)
      name: test name (required)
      module: 'reading' (default), 'listening', 'writing'
      difficulty: beginner / intermediate / advanced / expert (default intermediate)
      duration_minutes: int (default 60)
    """
    org = _check_center_admin(request.user, org_slug)

    upload = request.FILES.get('file')
    if not upload:
        return Response({'detail': 'No file uploaded.'}, status=400)
    if not upload.name.lower().endswith(('.xlsx', '.xlsm')):
        return Response({'detail': 'Please upload an .xlsx file.'}, status=400)

    name = (request.data.get('name') or '').strip()
    if not name:
        return Response({'detail': 'Test name is required.'}, status=400)

    module = (request.data.get('module') or 'reading').strip().lower()
    if module not in ('reading', 'listening', 'writing'):
        return Response({'detail': 'Module must be reading, listening, or writing.'}, status=400)

    difficulty = (request.data.get('difficulty') or 'intermediate').strip().lower()
    if difficulty not in ('beginner', 'intermediate', 'advanced', 'expert'):
        difficulty = 'intermediate'

    try:
        duration = int(request.data.get('duration_minutes') or 60)
    except (TypeError, ValueError):
        duration = 60
    duration = max(5, min(duration, 180))

    try:
        wb = openpyxl.load_workbook(upload, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return Response({'detail': f'Could not read the file: {exc}'}, status=400)

    headers, rows = _read_rows(wb)
    if not rows:
        return Response({'detail': 'The file has no data rows.'}, status=400)

    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        return Response(
            {'detail': f'Missing required columns: {", ".join(missing)}'},
            status=400,
        )

    # Validate every row first so we don't half-create a test on bad data
    errors = []
    for i, row in enumerate(rows, start=2):  # row 1 is the header
        passage_title = (str(row.get('Passage Title') or '').strip())
        if not passage_title:
            errors.append(f'Row {i}: Passage Title is empty.')
            continue
        qtype = str(row.get('Question Type') or '').strip().lower()
        if qtype not in VALID_TYPES:
            errors.append(
                f'Row {i}: Question Type "{row.get("Question Type")}" is invalid. '
                f'Must be one of: {", ".join(sorted(VALID_TYPES))}',
            )
        if not str(row.get('Question Text') or '').strip():
            errors.append(f'Row {i}: Question Text is empty.')
        if not str(row.get('Correct Answer') or '').strip():
            errors.append(f'Row {i}: Correct Answer is empty.')

    if errors:
        return Response({'detail': '; '.join(errors[:8])}, status=400)

    with transaction.atomic():
        test = Test.objects.create(
            organization=org,
            is_global=False,
            name=name,
            module=module,
            test_type='academic',
            difficulty=difficulty,
            duration_minutes=duration,
            description='Imported from Excel',
            is_published=True,
            status='published',
            created_by=request.user,
        )

        passages_by_title: dict[str, Passage] = {}
        passage_order = 0

        for row in rows:
            title = str(row.get('Passage Title') or '').strip()
            content = str(row.get('Passage Content') or '').strip()

            passage = passages_by_title.get(title)
            if passage is None:
                passage_order += 1
                passage = Passage.objects.create(
                    test=test,
                    part_number=passage_order,
                    title=title,
                    content=content if content and not content.startswith('(') else '',
                    order=passage_order,
                )
                passages_by_title[title] = passage
            elif content and not content.startswith('(') and not passage.content:
                # First non-empty content seen for this passage — fill it in
                passage.content = content
                passage.save(update_fields=['content'])

            qtype = str(row.get('Question Type')).strip().lower()
            options = None
            if qtype == 'mcq':
                options = [
                    str(row[c]).strip() for c in OPTION_COLUMNS
                    if row.get(c) is not None and str(row[c]).strip()
                ]
                options = options or None

            try:
                order = int(row.get('Question Number') or 0)
            except (TypeError, ValueError):
                order = 0

            Question.objects.create(
                passage=passage,
                question_type=qtype,
                text=str(row.get('Question Text')).strip(),
                options=options,
                correct_answer=str(row.get('Correct Answer')).strip(),
                order=order,
                points=1,
            )

    return Response({
        'success': True,
        'test_id': str(test.id),
        'name': test.name,
        'passages': len(passages_by_title),
        'questions': len(rows),
    }, status=201)
