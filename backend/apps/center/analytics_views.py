"""ETAP 7 — Markaz admini uchun analytics + Excel export.

Mock test natijalariga asoslangan statistika, oylik trend va eng yaxshi
talabalar ro'yxati. Export endpointi openpyxl orqali .xlsx fayl yaratadi.
"""

from datetime import timedelta
from io import BytesIO

from django.db.models import Avg, Count, Max, Min
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.mock.models import MockParticipant, MockSession
from apps.organizations.permissions import IsCenterAdmin

from .views import _OrgScopedViewSetMixin


class _CenterAdminBaseView(_OrgScopedViewSetMixin, APIView):
    """Slug → Organization, plus ensures the requester is a center admin."""

    permission_classes = [IsAuthenticated, IsCenterAdmin]
    kwargs: dict

    def initial(self, request, *args, **kwargs):
        self.kwargs = kwargs
        super().initial(request, *args, **kwargs)


SCORE_BUCKETS = [
    (0.0, 4.5, '< 4.5'),
    (4.5, 5.5, '4.5 – 5.5'),
    (5.5, 6.5, '5.5 – 6.5'),
    (6.5, 7.5, '6.5 – 7.5'),
    (7.5, 9.5, '7.5+'),
]


class CenterAnalyticsView(_CenterAdminBaseView):
    """GET /api/v1/center/<slug>/analytics/ — markaz statistikasi."""

    def get(self, request, **kwargs):
        org = self.get_organization()

        sessions = MockSession.objects.filter(organization=org)
        participants = MockParticipant.objects.filter(session__organization=org)
        completed = participants.filter(overall_band_score__isnull=False)

        students_count = org.users.filter(role='student').count()
        teachers_count = org.users.filter(role='teacher').count()

        agg = completed.aggregate(
            avg_overall=Avg('overall_band_score'),
            avg_listening=Avg('listening_score'),
            avg_reading=Avg('reading_score'),
            avg_writing=Avg('writing_score'),
            avg_speaking=Avg('speaking_score'),
            max_overall=Max('overall_band_score'),
            min_overall=Min('overall_band_score'),
        )

        def f(value):
            return round(float(value), 2) if value is not None else None

        score_distribution = []
        for low, high, label in SCORE_BUCKETS:
            count = completed.filter(
                overall_band_score__gte=low,
                overall_band_score__lt=high,
            ).count()
            score_distribution.append({'label': label, 'count': count})

        # Last 6 months trend
        now = timezone.now().date().replace(day=1)
        monthly = []
        for i in range(5, -1, -1):
            month_start = (now - timedelta(days=30 * i)).replace(day=1)
            next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
            month_qs = completed.filter(
                session__date__gte=month_start,
                session__date__lt=next_month,
            )
            avg = month_qs.aggregate(a=Avg('overall_band_score'))['a']
            monthly.append({
                'month': month_start.strftime('%b %Y'),
                'count': month_qs.count(),
                'avg_score': f(avg),
            })

        top_students = list(
            completed
            .order_by('-overall_band_score', '-session__date')
            .values(
                'id', 'full_name',
                'listening_score', 'reading_score',
                'writing_score', 'speaking_score', 'overall_band_score',
                'session__date', 'session__name',
            )[:10]
        )
        for row in top_students:
            for key in ('listening_score', 'reading_score', 'writing_score',
                        'speaking_score', 'overall_band_score'):
                if row[key] is not None:
                    row[key] = float(row[key])
            row['session_date'] = row.pop('session__date').isoformat()
            row['session_name'] = row.pop('session__name')

        recent_sessions = list(
            sessions.order_by('-date')[:8]
            .annotate(participants_total=Count('participants'))
            .values('id', 'name', 'date', 'status', 'participants_total')
        )
        for s in recent_sessions:
            s['date'] = s['date'].isoformat()

        return Response({
            'totals': {
                'students': students_count,
                'teachers': teachers_count,
                'sessions': sessions.count(),
                'participants': participants.count(),
                'completed': completed.count(),
            },
            'averages': {
                'overall': f(agg['avg_overall']),
                'listening': f(agg['avg_listening']),
                'reading': f(agg['avg_reading']),
                'writing': f(agg['avg_writing']),
                'speaking': f(agg['avg_speaking']),
                'max_overall': f(agg['max_overall']),
                'min_overall': f(agg['min_overall']),
            },
            'score_distribution': score_distribution,
            'monthly_trend': monthly,
            'top_students': top_students,
            'recent_sessions': recent_sessions,
        })


class CenterAnalyticsExcelView(_CenterAdminBaseView):
    """GET /api/v1/center/<slug>/analytics/export.xlsx — Excel export."""

    def get(self, request, **kwargs):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return Response(
                {'detail': 'openpyxl kutubxonasi o‘rnatilmagan.'},
                status=500,
            )

        org = self.get_organization()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Mock Results'

        headers = [
            '#', 'Talaba', 'Sessiya', 'Sana',
            'Listening', 'Reading', 'Writing', 'Speaking', 'Overall',
            'Writing status', 'Speaking status',
        ]
        header_fill = PatternFill('solid', fgColor='1F2937')
        header_font = Font(bold=True, color='FFFFFF')
        for col, value in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        rows = (
            MockParticipant.objects
            .filter(session__organization=org)
            .select_related('session')
            .order_by('-session__date', 'full_name')
        )

        for idx, p in enumerate(rows, start=1):
            ws.cell(row=idx + 1, column=1, value=idx)
            ws.cell(row=idx + 1, column=2, value=p.full_name)
            ws.cell(row=idx + 1, column=3, value=p.session.name)
            ws.cell(row=idx + 1, column=4, value=p.session.date.strftime('%d.%m.%Y'))
            for col, attr in enumerate(
                ('listening_score', 'reading_score', 'writing_score',
                 'speaking_score', 'overall_band_score'),
                start=5,
            ):
                value = getattr(p, attr)
                ws.cell(
                    row=idx + 1, column=col,
                    value=float(value) if value is not None else None,
                )
                ws.cell(row=idx + 1, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=idx + 1, column=10, value=p.get_writing_status_display())
            ws.cell(row=idx + 1, column=11, value=p.get_speaking_status_display())

        widths = [4, 28, 28, 12, 10, 10, 10, 10, 10, 14, 14]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # Summary
        last_row = rows.count() + 3
        completed = rows.filter(overall_band_score__isnull=False)
        summary = completed.aggregate(
            total=Count('id'), avg=Avg('overall_band_score'),
            best=Max('overall_band_score'),
        )
        ws.cell(row=last_row, column=1, value='XULOSA').font = Font(bold=True, size=12)
        ws.cell(row=last_row + 1, column=1, value='Tugatilgan testlar:')
        ws.cell(row=last_row + 1, column=2, value=summary['total'] or 0)
        ws.cell(row=last_row + 2, column=1, value='O‘rtacha Overall:')
        ws.cell(
            row=last_row + 2, column=2,
            value=round(float(summary['avg']), 2) if summary['avg'] else None,
        )
        ws.cell(row=last_row + 3, column=1, value='Eng yuqori:')
        ws.cell(
            row=last_row + 3, column=2,
            value=float(summary['best']) if summary['best'] else None,
        )

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        slug_safe = (org.slug or 'center').replace('/', '_')
        filename = f'IELTS_Mock_Results_{slug_safe}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
