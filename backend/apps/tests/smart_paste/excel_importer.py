"""ETAP 24 — Excel template importer.

Sheet 'Test' format (header row exact):
| section_kind | section_order | content_url | q_order | q_type | q_text | options | answer |

`options` cell format: "A:Foo|B:Bar|C:Baz".
`answer` for matching_headings groups: one row per paragraph (heading id in `answer`),
the q_text should be like "Paragraph B".
"""

import openpyxl
from openpyxl.workbook import Workbook

from .detector import DetectionResult
from .reading_parser import (
    ParsedQuestion,
    ParsedSection,
    ParseResult,
    build_answer_key,
    build_payload,
)


COLUMNS = [
    "section_kind", "section_order", "content_url",
    "q_order", "q_type", "q_text", "options", "answer",
]


def generate_template_xlsx(filepath: str) -> None:
    """Yozadi: rasmiy ILDIZmock Excel test shabloni."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    ws.append(COLUMNS)
    ws.append(["reading", 1, "", 1, "tfng",
               "The author claims X is true.", "", "TRUE"])
    ws.append(["reading", 1, "", 2, "mcq_single",
               "What is the capital?",
               "A:Paris|B:Berlin|C:Madrid|D:Rome", "A"])
    ws.append(["reading", 1, "", 3, "sentence_completion",
               "The first lesson starts on {{1}}.", "", "Monday"])
    ws.append([])
    ws.append([
        "# Use 'matching_headings' for groups; one row per paragraph,",
        "all rows in the same section_order.",
        "", "", "", "", "", "",
    ])
    wb.save(filepath)


def parse_excel(file_path: str) -> ParseResult:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "Test" not in wb.sheetnames:
        result = ParseResult(passage_html="", passage_word_count=0, paragraphs=[])
        result.errors.append("Workbook must have a 'Test' sheet.")
        return result

    ws = wb["Test"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or tuple(rows[0][:len(COLUMNS)]) != tuple(COLUMNS):
        result = ParseResult(passage_html="", passage_word_count=0, paragraphs=[])
        result.errors.append(f"First row must be: {' | '.join(COLUMNS)}")
        return result

    result = ParseResult(passage_html="", passage_word_count=0, paragraphs=[])
    sections_by_order: dict[int, ParsedSection] = {}

    for i, row in enumerate(rows[1:], start=2):
        if not row or not row[0] or (
            isinstance(row[0], str) and row[0].startswith("#")
        ):
            continue

        kind, sec_order, content_url, q_order, q_type, q_text, options, answer = row[:8]

        sec_key = int(sec_order) if sec_order else 1
        if sec_key not in sections_by_order:
            sections_by_order[sec_key] = ParsedSection(
                instructions=f"{(kind or 'reading').title()} Section {sec_key}",
            )
        sec = sections_by_order[sec_key]

        # "A:Foo|B:Bar"
        opts: list[tuple[str, str]] = []
        if options:
            for piece in str(options).split("|"):
                if ":" in piece:
                    a, b = piece.split(":", 1)
                    opts.append((a.strip(), b.strip()))

        qtype = (q_type or "").strip()
        try:
            q_order_int = int(q_order or 1)
        except (TypeError, ValueError):
            result.warnings.append(f"Row {i}: q_order is not a number — skipping")
            continue

        payload = build_payload(qtype, q_order_int, str(q_text or ""),
                                opts, str(q_text or ""))
        akey = build_answer_key(qtype, q_order_int, str(answer or ""))

        sec.questions.append(ParsedQuestion(
            order=q_order_int,
            qtype=qtype,
            payload=payload,
            answer_key=akey,
            detection=DetectionResult(qtype, 1.0, "explicit from Excel"),
            raw_text=str(q_text or ""),
        ))

    result.sections = [sections_by_order[k] for k in sorted(sections_by_order.keys())]
    if not result.sections:
        result.errors.append("No data rows found in the Excel sheet.")
    return result
